from django.http import HttpResponse
from django.shortcuts import render
from .import mapping
import openpyxl
from .import models
from datetime import datetime
import numpy as np
import pandas as pd
from pandas import DataFrame
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib import pylab
from pylab import *
import PIL, PIL.Image
import io
from io import StringIO
from io import BytesIO
import base64
from django.template.loader import get_template
from six.moves import urllib
import matplotlib.cm as cm
from matplotlib.ticker import MaxNLocator
from matplotlib.ticker import ScalarFormatter
import textwrap
import json
from django.shortcuts import render
from plotly.offline import plot
from plotly.graph_objs import Scatter
import plotly.graph_objects as go
import plotly.figure_factory as ff
import plotly.offline as opy
from plotly.graph_objs import Scatter, Layout, Figure, Data, Stream, YAxis, Marker
import chart_studio.plotly as py
NO = 0
TITLE = 1
INVENTORS = 2
APPLICANTS = 3
PUBLICATION_NUMBER = 4
COUNTRY = 5
EARLIEST_PRIORITY = 6
ipc = 7
cpc = 8
PUBLICATION_DATE = 9
PUBLICATION_YEAR = 10
EARLIEST_PUBLICATION = 11
FAMILY_NUMBER = 12

def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]


        # print('teekshan')
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active
        print(worksheet)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)



        rownum = 0
        for row in worksheet.iter_rows(values_only=True):

            # for cell in row:
             header = []
             if rownum == 0:
                header.append(row)
                rownum += 1
             else:


                  the_row = models.report( No =row[NO], Title=row[TITLE], Inventors=row[INVENTORS], Applicants=row[APPLICANTS],
                                         Publication_number=row[PUBLICATION_NUMBER], Country=row[COUNTRY], Earliest_priority=row[EARLIEST_PRIORITY],
                                         IPC=row[ipc],CPC=row[cpc], Publication_date=row[PUBLICATION_DATE], Publication_Year=row[PUBLICATION_YEAR],
                                         Earliest_publication=row[EARLIEST_PUBLICATION],Family_number=row[FAMILY_NUMBER]
                                         )
                  the_row.save()

        return render(request, 'index.html', {"excel_data": excel_data})


def dashboard1(request):

    data = models.report.objects.all()





    '''--------------------------------- COUNTRY CODE FROM PUBLICATION NO--------------------------------------------'''

    g1 = list()
    for row in data:
        g1.append(row.Publication_number)

    a1 = list()

    def first2(s):
        return s[:2]

    for i in g1:
        str = i
        two = first2(str)
        a1.append(two)

    b1 = np.array(a1)
    unique_elements1, counts_elements1 = np.unique(b1, return_counts=True)

    counts_elements1 = list(counts_elements1)
    unique_elements1 = list(unique_elements1)

    '''--------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX--------------------------------------------'''


    '''--------------------------------- FOR YEAR WISE PLOT USING PUBLICATION DATE-----------------------------------'''


    
    g2 = list()
    for row in data:
        g2.append(row.Publication_date)
    
    
    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)

    b2 = np.array(a2)
    unique_elements2, counts_elements2 = np.unique(b2, return_counts=True)

    counts_elements2 = list(counts_elements2)
    unique_elements2 = list(unique_elements2)

    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''

    '''--------------------------------- FOR YEAR WISE PLOT USING EARLIEST PRIORITY DATE-----------------------------'''


    g3 = list()
    for row in data:
        g3.append(row.Earliest_priority)
    
    
    a3 = list()
    

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a3 = list()
    for i in g3:
        datestring = i
        dt = try_parsing_date(datestring)
        a3.append(dt.year)

    b3 = np.array(a3)
    unique_elements3, counts_elements3 = np.unique(b3, return_counts=True)

    counts_elements3 = list(counts_elements3)
    unique_elements3 = list(unique_elements3)

    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''
    '''------------------------------- list of applicants - ---------------------------------------------------------'''

    g4 = list()
    for row in data:
        g4.append(row.Applicants)

    a4 = list()
    for i in g4:
        str = i
        a4.append(str.split('\n'))

    list_of_applicants = []

    for sublist in a4:
        for val in sublist:
            list_of_applicants.append(val)


    b4 = np.array(list_of_applicants)
    unique_elements4, counts_elements4 = np.unique(b4, return_counts=True)


    counts_elements4 = list(counts_elements4)
    unique_elements4 = list(unique_elements4)
    unique_elements4 = [x for _, x in sorted(zip(counts_elements4, unique_elements4))]

    counts_elements4.sort()

    def Reverse(lst):
        lst.reverse()
        return lst
    Reverse(counts_elements4)
    Reverse(unique_elements4)
    unique_elements4=unique_elements4[:10]
    counts_elements4=counts_elements4[:10]



    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''
    '''------------------------------- list of inventors - ----------------------------------------------------------'''

    g5 = list()
    for row in data:
        g5.append(row.Inventors)

    a5 = list()
    for i in g5:
        str = i
        a5.append(str.split('\n'))

    list_of_inventors = []

    for sublist in a5:
        for val in sublist:
            list_of_inventors.append(val)

    b5 = np.array(list_of_inventors)
    unique_elements5, counts_elements5 = np.unique(b5, return_counts=True)


    counts_elements5 = list(counts_elements5)
    unique_elements5 = list(unique_elements5)
    unique_elements5 = [x for _, x in sorted(zip(counts_elements5, unique_elements5))]

    counts_elements5.sort()

    def Reverse(lst):
        lst.reverse()
        return lst
    Reverse(counts_elements5)
    Reverse(unique_elements5)
    unique_elements5=unique_elements5[:10]
    counts_elements5=counts_elements5[:10]
    '''------------------------------- list of inventors - ----------------------------------------------------------'''
    '''------------------------------- Heat map for Publication year - ----------------------------------------------'''

    g11 = list()
    for row in data:
        g11.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g11:
        str = i
        two = first2(str)
        Country.append(two)


    g22 = list()
    for row in data:
        g22.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g22:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    zippedList = list(zip(Country, Year))
    dfObj = pd.DataFrame(zippedList, columns=['Country', 'Year'])
    dfObj.groupby(["Country", "Year"]).size().reset_index(name="count")
    df = pd.crosstab(dfObj.Year, dfObj.Country)

    ay = sns.heatmap(df, annot=True, cmap="coolwarm_r", fmt='g', linewidths=.5,
                     annot_kws={"style": "italic", "weight": "bold"})  # notation: "annot" not "annote"
    plt.title("Heatmap for Publication year and respective country")
    b, t = plt.ylim()  # discover the values for bottom and top
    b += 0.5  # Add 0.5 to the bottom
    t -= 0.5  # Subtract 0.5 from the top
    plt.ylim(b, t)  # update the ylim(bottom, top) values
    figure = plt.gcf()
    buf = io.BytesIO()
    figure.savefig(buf, format='png', transparent=True, quality=50, dpi=100)

    buf.seek(0)

    imsrc = base64.b64encode(buf.read())
    imuri = 'data:image/png;base64,{}'.format(urllib.parse.quote(imsrc))
    plt.close()
    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''

    '''------------------------------- Heat map for earliest priority - ---------------------------------------------'''
    g33 = list()
    for row in data:
        g33.append(row.Publication_number)

    Country_ep = list()

    def first2(s):
        return s[:2]

    for i in g33:
        str = i
        two = first2(str)
        Country_ep.append(two)

    g44 = list()
    for row in data:
        g44.append(row.Earliest_priority)


    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g44:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    zippedList = list(zip(Country_ep, Earliest_year))
    dfObj1 = pd.DataFrame(zippedList, columns=['Country_ep', 'Earliest_year'])
    dfObj1.groupby(["Country_ep", "Earliest_year"]).size().reset_index()

    df1 = pd.crosstab(dfObj1.Earliest_year, dfObj1.Country_ep)

    ay1 = sns.heatmap(df1, annot=True, cmap="coolwarm_r", fmt='g', linewidths=.5,
                     annot_kws={"style": "italic", "weight": "bold"})  # notation: "annot" not "annote"
    plt.title("Heatmap for Earliest priority year and respective country")
    b, t = plt.ylim()  # discover the values for bottom and top
    b += 0.5  # Add 0.5 to the bottom
    t -= 0.5  # Subtract 0.5 from the top
    plt.ylim(b, t)  # update the ylim(bottom, top) values

    figure1 = plt.gcf()
    buf1 = io.BytesIO()
    figure1.savefig(buf1, format='png', transparent=True, quality=50, dpi=100)
    buf1.seek(0)
    imsrc1 = base64.b64encode(buf1.read())
    imuri1 = 'data:image1/png;base64,{}'.format(urllib.parse.quote(imsrc1))

    plt.close()
    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''

    '''------------------------------- Bubble map for Publication year - --------------------------------------------'''

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)


    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)


    zippedList = list(zip(Country, Year))
    dfObj3 = pd.DataFrame(zippedList, columns=['Country', 'Year'])

    dfObj3 = dfObj3.groupby(['Country', 'Year']).size().to_frame(name='Count').reset_index()
    dfObj3['size1'] = dfObj3.apply(lambda row: (row.Count * 50), axis=1)

    N = 1
    ax = plt.figure().gca()
    ax = sns.scatterplot(dfObj3.Country, dfObj3.Year, alpha=0.5, s=dfObj3.size1)
    plt.title("Bubble map for Publication year and respective country")
    # plt.figure(figsize=(45, 40))
    for line in range(0, dfObj3.shape[0]):
        ax.text(dfObj3.Country[line], dfObj3.Year[line], dfObj3.Count[line], horizontalalignment='center',
                size='medium', color='black')
    colors = cm.rainbow(np.random.rand(N))
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))


    figure2 = plt.gcf()
    buf2 = io.BytesIO()
    figure2.savefig(buf2, format='png', transparent=True, quality=50, dpi=100)
    buf2.seek(0)
    imsrc2 = base64.b64encode(buf2.read())
    imuri2 = 'data:image1/png;base64,{}'.format(urllib.parse.quote(imsrc2))
    plt.close()
    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''

    '''------------------------------- Bubble map for Earliest priority year- ---------------------------------------'''

    g77 = list()
    for row in data:
        g77.append(row.Publication_number)

    Country_ep = list()

    def first2(s):
        return s[:2]

    for i in g77:
        str = i
        two = first2(str)
        Country_ep.append(two)

    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    zippedList = list(zip(Country_ep, Earliest_year))
    dfObj4 = pd.DataFrame(zippedList, columns=['Country_ep', 'Earliest_year'])


    dfObj4 = dfObj4.groupby(['Country_ep', 'Earliest_year']).size().to_frame(name='Count').reset_index()
    dfObj4['size2'] = dfObj4.apply(lambda row: (row.Count * 50), axis=1)

    N = 1
    ax = plt.figure().gca()
    ax = sns.scatterplot(dfObj4.Country_ep, dfObj4.Earliest_year, alpha=0.5,s=dfObj4.size2)
    plt.title("Bubble map for Earliest priority year and respective country")

    for line in range(0, dfObj4.shape[0]):
        ax.text(dfObj4.Country_ep[line], dfObj4.Earliest_year[line], dfObj4.Count[line], horizontalalignment='center',
                size='medium', color='black')
    colors = cm.rainbow(np.random.rand(N))
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))


    figure3 = plt.gcf()
    buf3 = io.BytesIO()
    figure3.savefig(buf3, format='png', transparent=True, quality=50, dpi=100)
    buf3.seek(0)
    imsrc3 = base64.b64encode(buf3.read())
    imuri3 = 'data:image1/png;base64,{}'.format(urllib.parse.quote(imsrc3))
    plt.close()
    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''

    '''-------------------------------Bubble map for top 5 inventors and priority year-------------------------------'''
    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventors = []

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)

    s88 = list()
    for row in data:
        s88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in s88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    b = list()
    c = 0
    for x in aa5:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0

    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Earliest_year[i])
        i = i + 1

    zippedList = list(zip(list_of_inventors, gk))
    dfObj44 = pd.DataFrame(zippedList, columns=['list_of_inventors', 'priority_year'])
    dfObj44 = dfObj44.groupby(['list_of_inventors', 'priority_year']).size().to_frame(name='Count').reset_index()
    dfObj44['size22'] = dfObj44.apply(lambda row: (row.Count * 1000), axis=1)
    df = dfObj44.nlargest(5, 'Count').reset_index()

    ax = plt.figure().gca()
    ax = sns.scatterplot( df.list_of_inventors,df.priority_year, alpha=0.5, s=df.size22,color='r')
    for line in range(0, df.shape[0]):
        ax.text( df.list_of_inventors[line],df.priority_year[line], df.Count[line], horizontalalignment='center',
                size='medium', color='black')
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))
    plt.xticks(rotation=45, ha='right')

    plt.title('Bubble map for top 5 inventors and priority year')
    plt.tight_layout()
    f = lambda y: textwrap.fill(y.get_text(), 20)
    ax.set_xticklabels(map(f, ax.get_xticklabels()))
    figure33 = plt.gcf()
    buf33 = io.BytesIO()
    figure33.savefig(buf33, format='png', transparent=True, quality=50, dpi=100)
    buf33.seek(0)
    imsrc33 = base64.b64encode(buf33.read())
    imuri33 = 'data:image1/png;base64,{}'.format(urllib.parse.quote(imsrc33))
    plt.close()
    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''

    '''-------------------------------Bubble map for top 5 applicants and publication year---------------------------'''
    g4 = list()
    for row in data:
        g4.append(row.Applicants)

    a4 = list()
    for i in g4:
        str = i
        a4.append(str.split('\n'))

    list_of_applicants = []

    for sublist in a4:
        for val in sublist:
            list_of_applicants.append(val)


    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    b = list()
    c = 0
    for x in a4:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0

    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Year[i])
        i = i + 1

    zippedList = list(zip(list_of_applicants, gk))
    dfObj444 = pd.DataFrame(zippedList, columns=['list_of_applicants', 'Year'])
    dfObj444 = dfObj444.groupby(['list_of_applicants', 'Year']).size().to_frame(name='Count1').reset_index()
    dfObj444['size222'] = dfObj444.apply(lambda row: (row.Count1 * 100), axis=1)
    df = dfObj444.nlargest(5, 'Count1').reset_index()


    ax1 = plt.figure().gca()

    ax1 = sns.scatterplot(df.list_of_applicants, df.Year, alpha=0.5, s=df.size222,color='r')

    for line in range(0, df.shape[0]):
        ax1.text(df.list_of_applicants[line], df.Year[line], df.Count1[line], horizontalalignment='center',
                size='medium', color='black')

    ax1.yaxis.set_major_locator(MaxNLocator(integer=True))
    plt.xticks(rotation=45, ha='right')
    plt.title('Bubble map for top 5 applicants and publication year')
    plt.tight_layout()
    f = lambda y: textwrap.fill(y.get_text(), 20)
    ax1.set_xticklabels(map(f, ax1.get_xticklabels()))

    figure333 = plt.gcf()
    buf333 = io.BytesIO()
    figure333.savefig(buf333, format='png', transparent=True, quality=50, dpi=100)
    buf333.seek(0)
    imsrc333 = base64.b64encode(buf333.read())
    imuri333 = 'data:image1/png;base64,{}'.format(urllib.parse.quote(imsrc333))
    plt.close()
    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''
    '''-------------------------------Bubble map for top 5 inventors and publication year----------------------------'''
    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventors = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    b = list()
    c = 0
    for x in aa5:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0

    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Year[i])
        i = i + 1




    zippedList = list(zip(list_of_inventors, gk))
    dfObj444 = pd.DataFrame(zippedList, columns=['list_of_inventors', 'Year'])
    dfObj444 = dfObj444.groupby(['list_of_inventors', 'Year']).size().to_frame(name='Count1').reset_index()
    dfObj444['size222'] = dfObj444.apply(lambda row: (row.Count1 * 1000), axis=1)

    df = dfObj444.nlargest(5, 'Count1').reset_index()


    ax1 = plt.figure().gca()


    ax1 = sns.scatterplot(df.list_of_inventors, df.Year, alpha=0.5, s=df.size222, color='r')

    for line in range(0, df.shape[0]):
        ax1.text(df.list_of_inventors[line], df.Year[line], df.Count1[line], horizontalalignment='center',
                 size='medium', color='black')
    plt.xticks(rotation=45, ha='right')
    plt.title('Bubble map for top 5 inventors and publication year')
    ax1.yaxis.set_major_locator(MaxNLocator(integer=True))
    plt.tight_layout()
    f = lambda y: textwrap.fill(y.get_text(), 20)
    ax1.set_xticklabels(map(f, ax1.get_xticklabels()))

    figure3333 = plt.gcf()
    buf3333 = io.BytesIO()
    figure3333.savefig(buf3333, format='png', transparent=True, quality=50, dpi=100)
    buf3333.seek(0)
    imsrc3333 = base64.b64encode(buf3333.read())
    imuri3333 = 'data:image1/png;base64,{}'.format(urllib.parse.quote(imsrc3333))
    plt.close()
    '''------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - -------------------------------------------'''
    '''-------------------------------Bubble map of top 5 applicants and priority------------------------------------'''
    g4 = list()
    for row in data:
        g4.append(row.Applicants)

    a4 = list()
    for i in g4:
        str = i
        a4.append(str.split('\n'))

    list_of_applicants = list()

    for sublist in a4:
        for val in sublist:
            list_of_applicants.append(val)

    s88 = list()
    for row in data:
        s88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in s88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    b = list()
    c = 0
    for x in a4:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0

    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Earliest_year[i])
        i = i + 1

    zippedList = list(zip(list_of_applicants, gk))
    dfObj44 = pd.DataFrame(zippedList, columns=['list_of_applicants', 'priority_year'])
    dfObj44 = dfObj44.groupby(['list_of_applicants', 'priority_year']).size().to_frame(name='Count').reset_index()
    dfObj44['size22'] = dfObj44.apply(lambda row: (row.Count * 100), axis=1)
    df = dfObj44.nlargest(5, 'Count').reset_index()

    ax = plt.figure().gca()
    ax = sns.scatterplot(df.list_of_applicants, df.priority_year, alpha=0.5, s=df.size22, color='r')
    for line in range(0, df.shape[0]):
        ax.text(df.list_of_applicants[line], df.priority_year[line], df.Count[line], horizontalalignment='center',
                size='medium', color='black')


    plt.xticks(rotation=45, ha='right')
    plt.title('Bubble map of top 5 applicants and priority')
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))
    plt.tight_layout()
    f = lambda y: textwrap.fill(y.get_text(), 20)
    ax.set_xticklabels(map(f, ax.get_xticklabels()))
    figure33333 = plt.gcf()
    buf33333 = io.BytesIO()
    figure33333.savefig(buf33333, format='png', transparent=True, quality=50, dpi=100)
    buf33333.seek(0)
    imsrc33333 = base64.b64encode(buf33333.read())
    imuri33333 = 'data:image1/png;base64,{}'.format(urllib.parse.quote(imsrc33333))
    plt.close()

    models.report.objects.all().delete()


    return render(request, 'charts.html',{"webdata5":unique_elements5,"wedata5":counts_elements5,"webdata1":unique_elements1,"wedata1":counts_elements1,"webdata2":unique_elements2,"wedata2":counts_elements2,"webdata3":unique_elements3,"wedata3":counts_elements3,"webdata4":unique_elements4,"wedata4":counts_elements4,'plot': imuri,'plot1': imuri1,'plot2':imuri2,'plot3':imuri3,'plot4':imuri33,'plot5':imuri333,'plot6':imuri3333,'plot7':imuri33333})




def home(request):
    data = models.report.objects.all()
    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)


    x = np.array(Country)
    z=(np.unique(x))

    zz = [10, 20, 30]
    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventors = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)
    x = np.array(list_of_inventors)
    zc = (np.unique(x))

    s5 = list()
    for row in data:
        s5.append(row.Applicants)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_applicantsx = list()

    for sublist in aa5:
        for val in sublist:
            list_of_applicantsx.append(val)
    yt = np.array(list_of_applicantsx)
    yct = (np.unique(yt))

    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)
    x11 = np.array(a2)
    zxy1 = (np.unique(x11))

    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    yx11 = np.array(Earliest_year)
    yzxy1 = (np.unique(yx11))
    VB = np.insert(yzxy1, np.arange(len(zxy1)), zxy1)
    VBT = (np.unique(VB))

    return render(request, 'home.html', {'z0t':VBT,'z': z,'z3':zz,'zs':zc,'zs1':yct})


def check(request):
    data = models.report.objects.all()
    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventorsx = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventorsx.append(val)
    xtt = np.array(list_of_inventorsx)
    zct = (np.unique(xtt))

    s5 = list()
    for row in data:
        s5.append(row.Applicants)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_applicantsx = list()

    for sublist in aa5:
        for val in sublist:
            list_of_applicantsx.append(val)
    xttc = np.array(list_of_applicantsx)
    zctc = (np.unique(xttc))

    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)
    x11 = np.array(a2)
    zxy1 = (np.unique(x11))
    max1=np.max(zxy1)




    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    yx11 = np.array(Earliest_year)
    yzxy1 = (np.unique(yx11))
    min1=np.min(yzxy1)

    VB=np.insert(yzxy1, np.arange(len(zxy1)), zxy1)
    VBT=(np.unique(VB))

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Countryxyz = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Countryxyz.append(two)
    x1 = np.array(Countryxyz)
    zxy = (np.unique(x1))
    zz=[10,20,30]


    datacou=list()
    datacou = request.GET.getlist('the-id')
    if len(datacou)==0:
        datacou=zxy
    else:
         datacou = datacou




    datay=list()
    datax=list()
    datay= request.GET.getlist('the-id0')

    if len(datay)==0:
        datay=min1
    else:
         datay = datay[0]
         datay = int(datay)

    datax = request.GET.getlist('the-id00')
    if len(datax)==0:
        datax=max1
    else:
         datax = datax[0]
         datax = int(datax)


    datacoutt = request.GET['the-id1']
    datacoutt = int(datacoutt)


    datacoui = list()
    datacoui = request.GET.getlist('the-id2')

    datacoua = list()
    datacoua = request.GET.getlist('the-id3')

    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)


    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    zippedList = list(zip(a2,Country))
    dfObj4e = pd.DataFrame(zippedList, columns=['year', 'Country'])
    df = dfObj4e[(dfObj4e['year'] >= datay) & (dfObj4e['year'] <= datax)]
    dfc = df[df['Country'].isin(datacou)]

    Year_list = dfc['year'].tolist()
    b2 = np.array(Year_list)
    unique_elements2, counts_elements2 = np.unique(b2, return_counts=True)
    zippedList = list(zip(counts_elements2, unique_elements2))
    dfO = pd.DataFrame(zippedList, columns=['Number', 'year' ])

    nn=dfO['Number'].tolist()

    if len(nn) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations=[]
    trace1 = go.Bar( x=dfO['year'],
                     y=dfO['Number'],
                     hovertext=dfO['year'],
                     hoverinfo="text",
                     text=dfO['Number'],
                     textposition='auto'
                    )
    data111 = [trace1]

    layout11 = go.Layout(title="PUBLICATION OF PATENTS YEAR WISE", xaxis={'title': 'YEAR', 'showgrid': False, 'tickformat': ',d'},
                         yaxis={'title': 'NO OF PATENTS', 'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    figure111 = go.Figure(data=data111,layout=layout11)
    figure111.update_layout(annotations=annotations)
    div111 = opy.plot(figure111, auto_open=False, output_type='div')



    g3 = list()
    for row in data:
        g3.append(row.Earliest_priority)

    a3 = list()

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a3 = list()
    for i in g3:
        datestring = i
        dt = try_parsing_date(datestring)
        a3.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    zippedList = list(zip(a3, Country))
    dfObj4e = pd.DataFrame(zippedList, columns=['year', 'Country'])
    df = dfObj4e[(dfObj4e['year'] >= datay) & (dfObj4e['year'] <= datax)]
    dfc = df[df['Country'].isin(datacou)]
    Year_list = dfc['year'].tolist()
    b3 = np.array(Year_list)
    unique_elements3, counts_elements3 = np.unique(b3, return_counts=True)
    zippedList = list(zip(counts_elements3, unique_elements3))
    dfO = pd.DataFrame(zippedList, columns=['Number', 'year'])
    nc=dfO['Number'].tolist()
    if len(nc)==0:
        annotations=[
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = []
    trace2 = go.Bar(x=dfO['year'],
        y=dfO['Number'],
                    hovertext=dfO['year'],
                    hoverinfo="text",
        text=dfO['Number'],

                    textposition='auto')


    data2 = [trace2]

    layout1 = go.Layout(title="EARLIEST YEAR WISE PUBLICATION", xaxis={'title': 'YEAR', 'showgrid': False,'tickformat': ',d'},
                        yaxis={'title': 'NO OF PATENTS', 'showgrid': False, 'tickformat': ',d'}, plot_bgcolor='rgb(252, 243, 207)')
    figure1 = go.Figure(data=data2, layout=layout1)
    figure1.update_layout(annotations=annotations)
    divb8 = opy.plot(figure1, auto_open=False, output_type='div')



    g4 = list()
    for row in data:
        g4.append(row.Applicants)

    a4 = list()
    for i in g4:
        str = i
        a4.append(str.split('\n'))

    list_of_applicants = list()

    for sublist in a4:
        for val in sublist:
            list_of_applicants.append(val)

    s88 = list()
    for row in data:
        s88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in s88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)
    b = list()
    c = 0
    for x in a4:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0
    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Earliest_year[i])
            cc.append(Country[i])
        i = i + 1

    zippedList = list(zip(list_of_applicants, gk,cc))
    dfObj44 = pd.DataFrame(zippedList, columns=['list_of_applicants', 'priority_year','Country'])
    dfObj44 = dfObj44.groupby(['list_of_applicants', 'priority_year','Country']).size().to_frame(name='Count').reset_index()
    dfObj44['size22'] = dfObj44.apply(lambda row: (row.Count * 100), axis=1)
    df3 = dfObj44[(dfObj44['priority_year'] >= datay) & (dfObj44['priority_year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoua) == 0:
         df = dfc.nlargest(datacoutt, 'Count').reset_index()
    else:
          df = dfc[dfc['list_of_applicants'].isin(datacoua)]

    applicant_list3 = df['list_of_applicants'].tolist()
    priority_list3 = df['priority_year'].tolist()
    co_list3 = df['Count'].tolist()
    applicant_el = list()
    def first10(s):
        return s[:10]

    for i in applicant_list3:
        str = i
        ten = first10(str)
        applicant_el.append(ten)

    n = len(applicant_el)
    for i in range(n):
        applicant_el[i] = applicant_el[i] + '....'
    if len(co_list3) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(co_list3) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(applicant_el, priority_list3, co_list3)
        ]



    trace1 = go.Bar(x=applicant_el,
                     #x=applicant_list3,
                    y=co_list3,
                    hovertext=applicant_list3,
                    hoverinfo="text",
                    text=co_list3,

                    textposition='auto'
                    )
    if len(co_list3) == 0:
        co_list3.append(0)
    trace2 = go.Scatter(x=applicant_el,
                        y=priority_list3,
                        hovertext=applicant_list3,
                        hoverinfo="text",
                           name='Bubble',
                           mode='markers',
                        visible=False,
                           marker=dict(
                               color=priority_list3,
                               size=co_list3,
                               sizemode='area',
                               sizeref=2. * max(co_list3) / (50. ** 2),
                               sizemin=6,

                               showscale=True

                           ))
    datap = [trace1, trace2]

    updatemenus = [
        dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected applicants",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected applicants priority year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations}]),

            ]),
            direction="down",

            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.16,
            yanchor="top"
        ),
    ]

    layout = dict(title='Top selected applicants', showlegend=False,
                  updatemenus=updatemenus,xaxis={ 'showgrid': False},
                         yaxis={ 'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datap, layout=layout)

    divb10 = opy.plot(fig, auto_open=False, output_type='div')

    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))


    list_of_inventors = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)

    s88 = list()
    for row in data:
        s88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in s88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    b = list()
    c = 0
    for x in aa5:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0
    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Earliest_year[i])
            cc.append(Country[i])
        i = i + 1

    zippedList = list(zip(list_of_inventors, gk,cc))
    dfObj44 = pd.DataFrame(zippedList, columns=['list_of_inventors', 'priority_year','Country'])
    dfObj44 = dfObj44.groupby(['list_of_inventors', 'priority_year','Country']).size().to_frame(name='Count').reset_index()
    dfObj44['size22'] = dfObj44.apply(lambda row: (row.Count * 1), axis=1)
    df3 = dfObj44[(dfObj44['priority_year'] >= datay) & (dfObj44['priority_year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoui) == 0:
         df = dfc.nlargest(datacoutt, 'Count').reset_index()
    else:
          df = dfc[dfc['list_of_inventors'].isin(datacoui)]

    inventors_list1 = df['list_of_inventors'].tolist()
    priority_list1 = df['priority_year'].tolist()
    cou_list1 = df['Count'].tolist()
    inventor_el = list()

    def first10(s):
        return s[:10]

    for i in inventors_list1:
        str = i
        ten = first10(str)
        inventor_el.append(ten)

    n = len(inventor_el)
    for i in range(n):
        inventor_el[i] = inventor_el[i] + '....'
    if len(cou_list1) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(cou_list1) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(inventor_el, priority_list1, cou_list1)
        ]
    trace1 = go.Bar(x=inventor_el,
                    y=cou_list1,
                    hovertext=inventors_list1,
                    hoverinfo="text",
                    text=cou_list1,
                    name='Bar',
                    # visible=False,
                    textposition='auto'
                    )
    if len(cou_list1) == 0:
        cou_list1.append(0)
    trace2 = go.Scatter(x=inventor_el,
                        y=priority_list1,
                        hovertext=inventors_list1,
                        hoverinfo="text",
                        name='Bubble',
                        mode='markers',
                        visible=False,
                        marker=dict(
                            color=priority_list1,
                            size=cou_list1,
                            sizemode='area',
                            sizeref=2. * max(cou_list1) / (50. ** 2),
                            sizemin=6,

                            showscale=True

                        ))
    dataq = [trace1, trace2]

    updatemenus = [
        dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected inventors",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected inventors priority year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations}]),

            ]),
            direction="down",

            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.15,
            yanchor="top"
        ),
    ]
    layout = dict(title='Top selected inventors', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False},
                  yaxis={'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=dataq, layout=layout)

    divb11 = opy.plot(fig, auto_open=False, output_type='div')

    g4 = list()
    for row in data:
        g4.append(row.Applicants)

    a4 = list()
    for i in g4:
        str = i
        a4.append(str.split('\n'))

    list_of_applicants = []

    for sublist in a4:
        for val in sublist:
            list_of_applicants.append(val)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)
    b = list()
    c = 0
    for x in a4:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0
    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Year[i])
            cc.append(Country[i])
        i = i + 1

    zippedList = list(zip(list_of_applicants, gk,cc))
    dfObj444 = pd.DataFrame(zippedList, columns=['list_of_applicants', 'Year','Country'])
    dfObj444 = dfObj444.groupby(['list_of_applicants', 'Year','Country']).size().to_frame(name='Count1').reset_index()
    dfObj444['size222'] = dfObj444.apply(lambda row: (row.Count1 * 100), axis=1)
    df3 = dfObj444[(dfObj444['Year'] >= datay) & (dfObj444['Year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoua) == 0:
        df = dfc.nlargest(datacoutt, 'Count1').reset_index()
    else:
        df = dfc[dfc['list_of_applicants'].isin(datacoua)]
    applicant_list1 = df['list_of_applicants'].tolist()
    year_list2 = df['Year'].tolist()
    co_list2 = df['Count1'].tolist()
    applicant_el = list()

    def first10(s):
        return s[:10]

    for i in applicant_list1:
        str = i
        ten = first10(str)
        applicant_el.append(ten)

    n = len(applicant_el)
    for i in range(n):
        applicant_el[i] = applicant_el[i] + '...'
    if len(co_list2) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(co_list2) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(applicant_el, year_list2, co_list2)
        ]
    trace1 = go.Bar(x=applicant_el,
                    y=co_list2,
                    hovertext=applicant_list1,
                    hoverinfo="text",
                    text=co_list2,
                    name='Bar',
                    # visible=False,
                    textposition='auto'
                    )
    if len(co_list2) == 0:
        co_list2.append(0)
    trace2 = go.Scatter(
        x=applicant_el,
        y=year_list2,
        hovertext=applicant_list1,
        hoverinfo="text",
        visible=False,
        mode='markers',
        marker=dict(
            color=year_list2,
            size=co_list2,
            sizemode='area',
            sizeref=2. * max(co_list2) / (50. ** 2),
            sizemin=6,
            showscale=True
        )

    )
    datar = [trace1, trace2]

    updatemenus = [
    dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected applicants",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected applicants publication year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations}]),

            ]),
        direction="down",

        showactive=True,
        x=0.1,
        xanchor="left",
        y=1.16,
        yanchor="top"
        ),
    ]
    layout = dict(title='Top selected applicants', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False},
                  yaxis={'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datar, layout=layout)
    divb12 = opy.plot(fig, auto_open=False, output_type='div')

    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventors = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)
    b = list()
    c = 0
    for x in aa5:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0

    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Year[i])
            cc.append(Country[i])
        i = i + 1


    zippedList = list(zip(list_of_inventors, gk,cc))
    dfObj444 = pd.DataFrame(zippedList, columns=['list_of_inventors', 'Year','Country'])
    dfObj444 = dfObj444.groupby(['list_of_inventors', 'Year','Country']).size().to_frame(name='Count1').reset_index()
    dfObj444['size222'] = dfObj444.apply(lambda row: (row.Count1 * 1000), axis=1)

    df3 = dfObj444[(dfObj444['Year'] >= datay) & (dfObj444['Year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoui) == 0:
        df = dfc.nlargest(datacoutt, 'Count1').reset_index()
    else:
        df = dfc[dfc['list_of_inventors'].isin(datacoui)]

    inventor_list = df['list_of_inventors'].tolist()
    year_list1 = df['Year'].tolist()
    co_list1 = df['Count1'].tolist()
    inventor_el = list()

    def first10(s):
        return s[:10]

    for i in inventor_list:
        str = i
        ten = first10(str)
        inventor_el.append(ten)

    n = len(inventor_el)
    for i in range(n):
        inventor_el[i] = inventor_el[i] + '...'
    if len(co_list1) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(co_list1) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(inventor_el, year_list1, co_list1)
        ]
    trace9 = go.Bar(x=inventor_el,
                    y=co_list1,
                    hovertext=inventor_list,
                    hoverinfo="text",
                    text=co_list1,
                    name='Bar',

                    textposition='auto'
                    )
    if len(co_list1) == 0:
        co_list1.append(0)
    trace10 = go.Scatter(
        x=inventor_el,
        y=year_list1,
        hovertext=inventor_list,
        hoverinfo="text",
        name='Bubble',
        visible=False,
        mode='markers',
        marker=dict(
            color=year_list1,
            size=co_list1,
            sizemode='area',
            sizeref=2. * max(co_list1) / (50. ** 2),
            sizemin=6,

            showscale=True
        )

    )
    datas = [trace9, trace10]
    updatemenus = [
        dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected inventors",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected inventors publication year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations}]),

            ]),
            direction="down",
            # pad={"r": 10, "t": 10},
            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.15,
            yanchor="top"
        )
    ]
    layout = dict(title='Top selected inventors', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False},
                  yaxis={'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datas, layout=layout)
    divb13 = opy.plot(fig, auto_open=False, output_type='div')

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    zippedList = list(zip(Country, Year))
    dfObj1x = pd.DataFrame(zippedList, columns=['Country', 'Year'])
    dfObj1x = dfObj1x.groupby(["Country", "Year"]).size().to_frame(name='Count1').reset_index()
    df = dfObj1x[(dfObj1x['Year'] >= datay) & (dfObj1x['Year'] <= datax)]
    dfc = df[df['Country'].isin(datacou)]
    df1 = dfc.pivot(index='Country', columns='Year', values='Count1')


    zx = dfc.pivot(index='Country', columns='Year', values='Count1').values

    zx[np.isnan(zx)] = 0
    axc = zx.tolist()

    yaxis = list(df1.index.values)
    xaxis = list(df1.columns)

    z11 = axc
    x11 = xaxis
    y11 = yaxis

    annotations = go.Annotations()
    for n, row in enumerate(z11):
        for m, val in enumerate(row):
            annotations.append(go.Annotation(text=z11[n][m], x=x11[m], y=y11[n],
                                             showarrow=False, font=dict(
                    color="black",
                    size=14
                )))
    if dfc.empty == True:
        trace00 = go.Bar(x=dfc['Year'],
                         y=dfc['Country'],
                         )
    else:
         trace00 = go.Heatmap(x=x11, y=y11, z=z11, colorscale='YlOrRd', showscale=False, name="heat", visible=False)
    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    zippedList = list(zip(Country, Year))
    dfObjp = pd.DataFrame(zippedList, columns=['Country', 'Year'])

    dfObjp = dfObjp.groupby(['Country', 'Year']).size().to_frame(name='Count').reset_index()
    dfObjp['size1'] = dfObjp.apply(lambda row: (row.Count * 1), axis=1)

    df1 = dfObjp[(dfObjp['Year'] >= datay) & (dfObjp['Year'] <= datax)]
    dfc = df1[df1['Country'].isin(datacou)]
    Country_list = dfc['Country'].tolist()
    Year_list = dfc['Year'].tolist()
    Count_list = dfc['Count'].tolist()
    size1_list = dfc['size1'].tolist()
    html_table1 = dfc.to_html(index=False)
    annotations1 = [
        dict(
            x=x,
            y=y,
            text='d' if x < 0 else z,  # Some conditional to define outliers

            showarrow=False,
            font=dict(
                color="black",
                size=14
            ),
            xanchor='center',  # Position of text relative to x axis (left/right/center)
            yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

        ) for x, y, z in zip(Year_list,Country_list,  Count_list)
    ]
    if len(Count_list) == 0:
        Count_list.append(0)
    trace11 = go.Scatter(
        x=Year_list,
        y=Country_list,
        name="bubble",
        mode='markers',
        marker=dict(
            color=Year_list,
            size=Count_list,
            sizemode='area',
            sizeref=3. * max(Count_list) / (30. ** 3),
            sizemin=6,

            showscale=True
        )

    )
    datase = [trace00, trace11]
    updatemenus = list([
        dict(
            active=1,
            buttons=list([
                dict(label="heat",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "No of patents of countries in respective years ",
                            'xaxis': {'title': 'Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations}]),
                dict(label="bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "No of patents of countries in respective years",
                            'xaxis': {'title': 'Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations1}]),

            ]),
            direction="down",
            # pad={"r": 10, "t": 10},
            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.15,
            yanchor="top"
        ),
    ])
    layout = dict(title='No of patents of countries in respective years', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False}, yaxis={'showgrid': False},
                  plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datase, layout=layout)
    divb14 = opy.plot(fig, auto_open=False, output_type='div')

    g33 = list()
    for row in data:
        g33.append(row.Publication_number)

    Country_ep = list()

    def first2(s):
        return s[:2]

    for i in g33:
        str = i
        two = first2(str)
        Country_ep.append(two)

    g44 = list()
    for row in data:
        g44.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g44:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    zippedList = list(zip(Country_ep, Earliest_year))
    dfObj1 = pd.DataFrame(zippedList, columns=['Country_ep', 'Earliest_year'])
    dfObj1 = dfObj1.groupby(["Country_ep", "Earliest_year"]).size().to_frame(name='Count1').reset_index()
    df3 = dfObj1[(dfObj1['Earliest_year'] >= datay) & (dfObj1['Earliest_year'] <= datax)]
    dfc = df3[df3['Country_ep'].isin(datacou)]
    df4 = dfc.pivot(index='Country_ep', columns='Earliest_year', values='Count1')
    z = dfc.pivot(index='Country_ep', columns='Earliest_year', values='Count1').values
    z[np.isnan(z)] = 0

    a = z.tolist()
    yaxis1 = list(df4.index.values)
    xaxis1 = list(df4.columns)
    z111 = a
    x111 = xaxis1
    y111 = yaxis1
    annotations = go.Annotations()
    for n, row in enumerate(z111):
        for m, val in enumerate(row):
            annotations.append(go.Annotation(text=z111[n][m], x=x111[m], y=y111[n],
                                             showarrow=False, font=dict(
                    color="black",
                    size=14
                )))
    if dfc.empty == True:
        trace000 = go.Bar(x=dfc['Earliest_year'],
                         y=dfc['Country_ep'],
                         )
    else:
        trace000 = go.Heatmap(x=x111, y=y111, z=z111, colorscale='YlOrRd', showscale=False, name="heat", visible=False)
    g77 = list()
    for row in data:
        g77.append(row.Publication_number)

    Country_ep = list()

    def first2(s):
        return s[:2]

    for i in g77:
        str = i
        two = first2(str)
        Country_ep.append(two)

    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    zippedList = list(zip(Country_ep, Earliest_year))
    dfObj4e = pd.DataFrame(zippedList, columns=['Country_ep', 'Earliest_year'])

    dfObj4e = dfObj4e.groupby(['Country_ep', 'Earliest_year']).size().to_frame(name='Count').reset_index()
    dfObj4e['size2'] = dfObj4e.apply(lambda row: (row.Count * 1), axis=1)
    df2 = dfObj4e[(dfObj4e['Earliest_year'] >= datay) & (dfObj4e['Earliest_year'] <= datax)]
    dfc = df2[df2['Country_ep'].isin(datacou)]

    Country_liste = dfc['Country_ep'].tolist()
    Year_liste = dfc['Earliest_year'].tolist()
    Count_lise = dfc['Count'].tolist()
    size1_liste = dfc['size2'].tolist()
    pd.set_option('display.width', 2000)
    html_table = dfc.to_html(index=False)


    annotations1 = [
        dict(
            x=x,
            y=y,
            text='d' if x < 0 else z,  # Some conditional to define outliers

            showarrow=False,
            font=dict(
                color="black",
                size=14
            ),
            xanchor='center',  # Position of text relative to x axis (left/right/center)
            yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

        ) for x, y, z in zip(Year_liste,Country_liste,  Count_lise)
    ]
    if len(Count_lise) == 0:
        Count_lise.append(0)
    trace111 = go.Scatter(
        x=Year_liste,
        y=Country_liste,
        name="bubble",
        mode='markers',
        marker=dict(
            color=Year_liste,
            size=Count_lise,
            sizemode='area',
            sizeref=3. * max(Count_lise) / (30. ** 3),
            sizemin=6,

            showscale=True
        )

    )
    dataset = [trace000, trace111]
    updatemenus = list([
        dict(
            active=1,
            buttons=list([
                dict(label="heat",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "No of patents of countries in respective priority years ",
                            'xaxis': {'title': 'priority Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations}]),
                dict(label="bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "No of patents of countries in respective priority years",
                            'xaxis': {'title': 'priority Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations1}]),

            ]),
            direction="down",

            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.15,
            yanchor="top"
        ),
    ])
    layout = dict(title='No of patents of countries in respective priority years', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False}, yaxis={'showgrid': False},
                  plot_bgcolor='rgb(252, 243, 207)')

    fig = go.Figure(data=dataset, layout=layout)
    
    divb15 = opy.plot(fig, auto_open=False, output_type='div')

    number= list()
    for row in data:
        number.append(row.No)

    title=list()
    for row in data:
        title.append(row.Title)

    inventors=list()
    for row in data:
        inventors.append(row.Inventors)

    aa5 = list()
    for i in inventors:
        str = i
        aa5.append(str.split('\n'))



    applicants=list()
    for row in data:
        applicants.append(row.Applicants)
    zyz = list()
    for i in applicants:
        str = i
        zyz.append(str.split('\n'))



    pno = list()
    for row in data:
        pno.append(row.Publication_number)

    country=list()
    for row in data:
        country.append(row.Country)

    ep=list()
    for row in data:
        ep.append(row.Earliest_priority)

    fno=list()
    for row in data:
        fno.append(row.Family_number)

    ipc=list()
    for row in data:
       ipc.append(row.IPC)

    xx = list()
    for i in ipc:
        str = i
        xx.append(str.split('\n'))




    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Countryx = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Countryx.append(two)
    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    zippedList = list(zip(a2,Earliest_year, Countryx,number,title,aa5,zyz,pno,country,ep,fno,xx))
    df = pd.DataFrame(zippedList, columns=['yd','ed','cd','Number','Title','Inventors','Applicants','PNO','Country','Earliest_Priority','Family_No','IPC'])
    df0=df[(df['ed'] >= datay) & (df['ed'] <= datax)]

    dfc = df0[df0['cd'].isin(datacou)]
    '''if len(datacoui) == 0:
        datacoui = zct
    else:
        datacoui = datacoui

    if len(datacoua) == 0:
        datacoua = zctc
    else:
        datacoua = datacoua'''

    # df1 = df[pd.DataFrame(df.lists.values.tolist()).eq(c.values).any(axis=1)]
    '''dff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
    mask = dff.isin(datacoui).any(axis=1)
    dfr = dfc[mask]

    dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
    mask3 = dffff.isin(datacoua).any(axis=1)
    dfs = dfc[mask3]
    frames = [dfr, dfs]
    result = pd.concat(frames)
    dfzz = result.drop_duplicates(subset=['Number'], keep="first")'''
    #dfxy = dfc[dfc['Inventors'].isin(datacoui)]
    if (len(datacoui) != 0 and  len(datacoua) == 0):
        dfff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
        mask1 = dfff.isin(datacoui).any(axis=1)
        dfz = dfc[mask1]

    if (len(datacoui) == 0 and len(datacoua) != 0):
        dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
        mask2 = dffff.isin(datacoua).any(axis=1)
        dfz = dfc[mask2]

    if (len(datacoui) == 0 and len(datacoua) == 0):
        datacoua = zctc
        datacoui = zct
        dfff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
        mask1 = dfff.isin(datacoui).any(axis=1)
        dfx = dfc[mask1]

        dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
        mask2 = dffff.isin(datacoua).any(axis=1)
        dfy = dfc[mask2]
        frames = [dfx, dfy]
        result = pd.concat(frames)
        dfz = result.drop_duplicates(subset=['Number'], keep="first")

    if (len(datacoui) != 0 and len(datacoua) != 0):
        dfff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
        mask1 = dfff.isin(datacoui).any(axis=1)
        dfx = dfc[mask1]

        dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
        mask2 = dffff.isin(datacoua).any(axis=1)
        dfy = dfc[mask2]
        frames = [dfx, dfy]
        result = pd.concat(frames)
        dfz = result.drop_duplicates(subset=['Number'], keep="first")

    dfct=dfz.drop(['yd', 'ed','cd'], axis=1)
    dfct.sort_values(by=['Number'])
    html_tablex = dfct.to_html(index=False)
    request.session['html_tablex'] = html_tablex


    VBTl = VBT.tolist()
    zxyl = zxy.tolist()

    zctl = zct.tolist()
    zctcl = zctc.tolist()
    request.session['VBTl'] = VBTl
    request.session['zxyl'] = zxyl
    request.session['zz'] = zz
    request.session['zctl'] = zctl
    request.session['zctcl'] = zctcl
    request.session['div111'] = div111
    request.session['divb8'] = divb8
    request.session['divb10'] = divb10
    request.session['divb11'] = divb11
    request.session['divb12'] = divb12
    request.session['divb13'] = divb13
    request.session['divb14'] = divb14
    request.session['divb15'] = divb15

    return render(request, 'home.html', {'z0t':VBT,'z':zxy,'z3':zz,'zs':zct,'zs1':zctc,'plotdiv': div111,'plotdiv9': divb8,'plotdiv11': divb10,'plotdiv12': divb11,'plotdiv13': divb12,'plotdiv14': divb13,'plotdiv15': divb14,'plotdiv16': divb15,'html_table':html_table,'html_table1':html_table1})

def check1(request):
    data = models.report.objects.all()
    number = list()
    for row in data:
        number.append(row.No)

    title = list()
    for row in data:
        title.append(row.Title)

    inventors = list()
    for row in data:
        inventors.append(row.Inventors)

    aa5 = list()
    for i in inventors:
        str = i
        aa5.append(str.split('\n'))

    applicants = list()
    for row in data:
        applicants.append(row.Applicants)
    zyz = list()
    for i in applicants:
        str = i
        zyz.append(str.split('\n'))
    pno = list()
    for row in data:
        pno.append(row.Publication_number)

    country = list()
    for row in data:
        country.append(row.Country)

    ep = list()
    for row in data:
        ep.append(row.Earliest_priority)

    fno = list()
    for row in data:
        fno.append(row.Family_number)

    ipc = list()
    for row in data:
        ipc.append(row.IPC)

    xx = list()
    for i in ipc:
        str = i
        xx.append(str.split('\n'))

    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Countryx = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Countryx.append(two)
    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year1 = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year1.append(dt.year)


    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventorsx = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventorsx.append(val)
    xtt = np.array(list_of_inventorsx)
    zct = (np.unique(xtt))

    s5 = list()
    for row in data:
        s5.append(row.Applicants)

    aa5a = list()
    for i in s5:
        str = i
        aa5a.append(str.split('\n'))

    list_of_applicantsx = list()

    for sublist in aa5a:
        for val in sublist:
            list_of_applicantsx.append(val)
    xttc = np.array(list_of_applicantsx)
    zctc = (np.unique(xttc))

    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)
    x11 = np.array(a2)
    zxy1 = (np.unique(x11))
    max1 = np.max(zxy1)


    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    yx11 = np.array(Earliest_year)
    yzxy1 = (np.unique(yx11))
    min1 = np.min(yzxy1)

    VB = np.insert(yzxy1, np.arange(len(zxy1)), zxy1)
    VBT = (np.unique(VB))

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Countryxyz = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Countryxyz.append(two)
    x1 = np.array(Countryxyz)
    zxy = (np.unique(x1))
    zz = [10, 20, 30]

    datacou = list()
    datacou = request.GET.getlist('the-id')
    if len(datacou) == 0:
        datacou = zxy
    else:
        datacou = datacou



    datay = ""
    datax = ""
    datay = request.GET.getlist('the-id0')

    if len(datay) == 0:
        datay = min1
    else:
        datay = datay[0]
        datay = int(datay)

    datax = request.GET.getlist('the-id00')
    if len(datax) == 0:
        datax = max1
    else:
        datax = datax[0]
        datax = int(datax)


    datacoutt = request.GET['the-id1']
    datacoutt = int(datacoutt)

    datacoui = list()
    datacoui = request.GET.getlist('the-id2')

    datacoua = list()
    datacoua = request.GET.getlist('the-id3')
    '''if len(datacoui) == 0:
        datacoui = zct
    else:
        datacoui = datacoui
    datacoua = list()
    datacoua = request.GET.getlist('the-id3')
    if len(datacoua) == 0:
        datacoua = zctc
    else:
        datacoua = datacoua'''

    zippedList = list(zip(a2, Earliest_year1, Countryx, number, title, aa5, zyz, pno, country, ep, fno, xx))
    df = pd.DataFrame(zippedList,
                      columns=['yd', 'ed', 'cd', 'Number', 'Title', 'Inventors', 'Applicants', 'PNO', 'Country',
                               'Earliest_Priority', 'Family_No', 'IPC'])
    df0 = df[(df['ed'] >= datay) & (df['ed'] <= datax)]

    dfc = df0[df0['cd'].isin(datacou)]
    if (len(datacoui) != 0 and  len(datacoua) == 0):
        dfff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
        mask1 = dfff.isin(datacoui).any(axis=1)
        dfz = dfc[mask1]

    if (len(datacoui) == 0 and len(datacoua) != 0):
        dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
        mask2 = dffff.isin(datacoua).any(axis=1)
        dfz = dfc[mask2]

    if (len(datacoui) == 0 and len(datacoua) == 0):
        datacoua = zctc
        datacoui = zct
        dfff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
        mask1 = dfff.isin(datacoui).any(axis=1)
        dfx = dfc[mask1]

        dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
        mask2 = dffff.isin(datacoua).any(axis=1)
        dfy = dfc[mask2]
        frames = [dfx, dfy]
        result = pd.concat(frames)
        dfz = result.drop_duplicates(subset=['Number'], keep="first")


    if (len(datacoui) != 0 and len(datacoua) != 0):
        dfff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
        mask1 = dfff.isin(datacoui).any(axis=1)
        dfx = dfc[mask1]

        dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
        mask2 = dffff.isin(datacoua).any(axis=1)
        dfy = dfc[mask2]
        frames = [dfx, dfy]
        result = pd.concat(frames)
        dfz = result.drop_duplicates(subset=['Number'], keep="first")

    '''dfff = pd.DataFrame(dfc['Inventors'].values.tolist(), index=dfc.index)
    mask1 = dfff.isin(datacoui).any(axis=1)
    dfx = dfc[mask1]
  
    dffff = pd.DataFrame(dfc['Applicants'].values.tolist(), index=dfc.index)
    mask2 = dffff.isin(datacoua).any(axis=1)
    dfy = dfc[mask2]
    frames = [dfx, dfy]
    result = pd.concat(frames)
    dfz = result.drop_duplicates(subset=['Number'], keep="first")'''

    dfct = dfz.drop(['yd', 'ed', 'cd'], axis=1)
    dfct.sort_values(by=['Number'])
    html_tablex = dfct.to_html(index=False)

    '''VBTl = request.session['VBTl']
    zxyl = request.session['zxyl']
    zz = request.session['zz']
    zctl = request.session['zctl']
    zctcl = request.session['zctcl']'''
    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)


    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    zippedList = list(zip(a2,Country))
    dfObj4e = pd.DataFrame(zippedList, columns=['year', 'Country'])
    df = dfObj4e[(dfObj4e['year'] >= datay) & (dfObj4e['year'] <= datax)]
    dfc = df[df['Country'].isin(datacou)]

    Year_list = dfc['year'].tolist()
    b2 = np.array(Year_list)
    unique_elements2, counts_elements2 = np.unique(b2, return_counts=True)
    zippedList = list(zip(counts_elements2, unique_elements2))
    dfO = pd.DataFrame(zippedList, columns=['Number', 'year' ])

    nn=dfO['Number'].tolist()

    if len(nn) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations=[]
    trace1 = go.Bar( x=dfO['year'],
        y=dfO['Number'],
                     hovertext=dfO['year'],
                     hoverinfo="text",
        text=dfO['Number'],
                    textposition='auto'
                    )
    data111 = [trace1]

    layout11 = go.Layout(title="PUBLICATION OF PATENTS YEAR WISE", xaxis={'title': 'YEAR', 'showgrid': False, 'tickformat': ',d'},
                         yaxis={'title': 'NO OF PATENTS', 'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    figure111 = go.Figure(data=data111,layout=layout11)
    figure111.update_layout(annotations=annotations)
    div111 = opy.plot(figure111, auto_open=False, output_type='div')


    g3 = list()
    for row in data:
        g3.append(row.Earliest_priority)

    a3 = list()

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a3 = list()
    for i in g3:
        datestring = i
        dt = try_parsing_date(datestring)
        a3.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    zippedList = list(zip(a3, Country))
    dfObj4e = pd.DataFrame(zippedList, columns=['year', 'Country'])
    df = dfObj4e[(dfObj4e['year'] >= datay) & (dfObj4e['year'] <= datax)]
    dfc = df[df['Country'].isin(datacou)]
    Year_list = dfc['year'].tolist()
    b3 = np.array(Year_list)
    unique_elements3, counts_elements3 = np.unique(b3, return_counts=True)
    zippedList = list(zip(counts_elements3, unique_elements3))
    dfO = pd.DataFrame(zippedList, columns=['Number', 'year'])
    nc=dfO['Number'].tolist()
    if len(nc)==0:
        annotations=[
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = []
    trace2 = go.Bar(x=dfO['year'],
        y=dfO['Number'],
                    hovertext=dfO['year'],
                    hoverinfo="text",
        text=dfO['Number'],

                    textposition='auto')


    data2 = [trace2]

    layout1 = go.Layout(title="EARLIEST YEAR WISE PUBLICATION", xaxis={'title': 'YEAR', 'showgrid': False,'tickformat': ',d'},
                        yaxis={'title': 'NO OF PATENTS', 'showgrid': False, 'tickformat': ',d'}, plot_bgcolor='rgb(252, 243, 207)')
    figure1 = go.Figure(data=data2, layout=layout1)
    figure1.update_layout(annotations=annotations)
    divb8 = opy.plot(figure1, auto_open=False, output_type='div')


    g4 = list()
    for row in data:
        g4.append(row.Applicants)

    a4 = list()
    for i in g4:
        str = i
        a4.append(str.split('\n'))

    list_of_applicants = list()

    for sublist in a4:
        for val in sublist:
            list_of_applicants.append(val)

    s88 = list()
    for row in data:
        s88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in s88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)
    b = list()
    c = 0
    for x in a4:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0
    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Earliest_year[i])
            cc.append(Country[i])
        i = i + 1

    zippedList = list(zip(list_of_applicants, gk,cc))
    dfObj44 = pd.DataFrame(zippedList, columns=['list_of_applicants', 'priority_year','Country'])
    dfObj44 = dfObj44.groupby(['list_of_applicants', 'priority_year','Country']).size().to_frame(name='Count').reset_index()
    dfObj44['size22'] = dfObj44.apply(lambda row: (row.Count * 100), axis=1)
    df3 = dfObj44[(dfObj44['priority_year'] >= datay) & (dfObj44['priority_year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoua) == 0:
         df = dfc.nlargest(datacoutt, 'Count').reset_index()
    else:
          df = dfc[dfc['list_of_applicants'].isin(datacoua)]

    applicant_list3 = df['list_of_applicants'].tolist()
    priority_list3 = df['priority_year'].tolist()
    co_list3 = df['Count'].tolist()
    applicant_el = list()

    def first10(s):
        return s[:10]

    for i in applicant_list3:
        str = i
        ten = first10(str)
        applicant_el.append(ten)

    n = len(applicant_el)
    for i in range(n):
        applicant_el[i] = applicant_el[i] + '...'
    if len(co_list3) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(co_list3) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(applicant_el, priority_list3, co_list3)
        ]
    trace1 = go.Bar(x=applicant_el,
                    y=co_list3,
                    hovertext=applicant_list3,
                    hoverinfo="text",
                    text=co_list3,

                    textposition='auto'
                    )
    if len(co_list3) == 0:
        co_list3.append(0)
    trace2 = go.Scatter(x=applicant_el,
                           y=priority_list3,
                        hovertext=applicant_list3,
                        hoverinfo="text",
                        name='Bubble',
                        mode='markers',
                        visible=False,
                           marker=dict(
                               color=priority_list3,
                               size=co_list3,
                               sizemode='area',
                               sizeref=2. * max(co_list3) / (50. ** 2),
                               sizemin=6,

                               showscale=True

                           ))
    datap = [trace1, trace2]

    updatemenus = [
        dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected applicants",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected applicants priority year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations}]),

            ]),
            direction="down",

            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.16,
            yanchor="top"
        ),
    ]

    layout = dict(title='Top selected applicants', showlegend=False,
                  updatemenus=updatemenus,xaxis={ 'showgrid': False},
                         yaxis={ 'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datap, layout=layout)

    divb10 = opy.plot(fig, auto_open=False, output_type='div')

    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))


    list_of_inventors = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)

    s88 = list()
    for row in data:
        s88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in s88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    b = list()
    c = 0
    for x in aa5:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0
    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Earliest_year[i])
            cc.append(Country[i])
        i = i + 1

    zippedList = list(zip(list_of_inventors, gk,cc))
    dfObj44 = pd.DataFrame(zippedList, columns=['list_of_inventors', 'priority_year','Country'])
    dfObj44 = dfObj44.groupby(['list_of_inventors', 'priority_year','Country']).size().to_frame(name='Count').reset_index()
    dfObj44['size22'] = dfObj44.apply(lambda row: (row.Count * 1), axis=1)
    df3 = dfObj44[(dfObj44['priority_year'] >= datay) & (dfObj44['priority_year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoui) == 0:
         df = dfc.nlargest(datacoutt, 'Count').reset_index()
    else:
          df = dfc[dfc['list_of_inventors'].isin(datacoui)]

    inventors_list1 = df['list_of_inventors'].tolist()
    priority_list1 = df['priority_year'].tolist()
    cou_list1 = df['Count'].tolist()
    inventor_el = list()

    def first10(s):
        return s[:10]

    for i in inventors_list1:
        str = i
        ten = first10(str)
        inventor_el.append(ten)

    n = len(inventor_el)
    for i in range(n):
        inventor_el[i] = inventor_el[i] + '...'
    if len(cou_list1) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(cou_list1) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(inventors_list1, priority_list1, cou_list1)
        ]
    trace1 = go.Bar(x=inventor_el,
                    y=cou_list1,
                    hovertext=inventors_list1,
                    hoverinfo="text",
                    text=cou_list1,
                    name='Bar',
                    # visible=False,
                    textposition='auto'
                    )
    if len(cou_list1) == 0:
        cou_list1.append(0)
    trace2 = go.Scatter(x=inventor_el,
                        y=priority_list1,
                        hovertext=inventors_list1,
                        hoverinfo="text",
                        name='Bubble',
                        mode='markers',
                        visible=False,
                        marker=dict(
                            color=priority_list1,
                            size=cou_list1,
                            sizemode='area',
                            sizeref=2. * max(cou_list1) / (50. ** 2),
                            sizemin=6,

                            showscale=True

                        ))
    dataq = [trace1, trace2]

    updatemenus = [
        dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected inventors",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected inventors priority year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations}]),

            ]),
            direction="down",

            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.15,
            yanchor="top"
        ),
    ]
    layout = dict(title='Top selected inventors', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False},
                  yaxis={'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=dataq, layout=layout)

    divb11 = opy.plot(fig, auto_open=False, output_type='div')

    g4 = list()
    for row in data:
        g4.append(row.Applicants)

    a4 = list()
    for i in g4:
        str = i
        a4.append(str.split('\n'))

    list_of_applicants = []

    for sublist in a4:
        for val in sublist:
            list_of_applicants.append(val)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)
    b = list()
    c = 0
    for x in a4:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0
    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Year[i])
            cc.append(Country[i])
        i = i + 1

    zippedList = list(zip(list_of_applicants, gk,cc))
    dfObj444 = pd.DataFrame(zippedList, columns=['list_of_applicants', 'Year','Country'])
    dfObj444 = dfObj444.groupby(['list_of_applicants', 'Year','Country']).size().to_frame(name='Count1').reset_index()
    dfObj444['size222'] = dfObj444.apply(lambda row: (row.Count1 * 100), axis=1)
    df3 = dfObj444[(dfObj444['Year'] >= datay) & (dfObj444['Year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoua) == 0:
        df = dfc.nlargest(datacoutt , 'Count1').reset_index()
    else:
        df = dfc[dfc['list_of_applicants'].isin(datacoua)]
    applicant_list1 = df['list_of_applicants'].tolist()
    year_list2 = df['Year'].tolist()
    co_list2 = df['Count1'].tolist()
    applicant_el = list()

    def first10(s):
        return s[:10]

    for i in applicant_list1:
        str = i
        ten = first10(str)
        applicant_el.append(ten)

    n = len(applicant_el)
    for i in range(n):
        applicant_el[i] = applicant_el[i] + '...'
    if len(co_list2) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(co_list2) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(applicant_el, year_list2, co_list2)
        ]
    trace1 = go.Bar(x=applicant_el,
                    y=co_list2,
                    hovertext=applicant_list1,
                    hoverinfo="text",
                    text=co_list2,
                    name='Bar',
                    # visible=False,
                    textposition='auto'
                    )
    if len(co_list2) == 0:
        co_list2.append(0)
    trace2 = go.Scatter(
        x=applicant_el,
        y=year_list2,
        hovertext=applicant_list1,
        hoverinfo="text",
        visible=False,
        mode='markers',
        marker=dict(
            color=year_list2,
            size=co_list2,
            sizemode='area',
            sizeref=2. * max(co_list2) / (50. ** 2),
            sizemin=6,
            showscale=True
        )

    )
    datar = [trace1, trace2]

    updatemenus = [
    dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected applicants",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected applicants publication year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Applicants'},
                            "annotations": annotations}]),

            ]),
        direction="down",

        showactive=True,
        x=0.1,
        xanchor="left",
        y=1.16,
        yanchor="top"
        ),
    ]
    layout = dict(title='Top selected applicants', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False},
                  yaxis={'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datar, layout=layout)
    divb12 = opy.plot(fig, auto_open=False, output_type='div')

    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventors = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)
    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)
    b = list()
    c = 0
    for x in aa5:
        for y in x:
            c = c + 1
        b.append(c)
        c = 0

    cc = list()
    gk = list()
    i = 0
    for y in b:
        a = y
        for z in range(a):
            gk.append(Year[i])
            cc.append(Country[i])
        i = i + 1

    zippedList = list(zip(list_of_inventors, gk,cc))
    dfObj444 = pd.DataFrame(zippedList, columns=['list_of_inventors', 'Year','Country'])
    dfObj444 = dfObj444.groupby(['list_of_inventors', 'Year','Country']).size().to_frame(name='Count1').reset_index()
    dfObj444['size222'] = dfObj444.apply(lambda row: (row.Count1 * 1000), axis=1)

    df3 = dfObj444[(dfObj444['Year'] >= datay) & (dfObj444['Year'] <= datax)]
    dfc = df3[df3['Country'].isin(datacou)]
    if len(datacoui) == 0:
        df = dfc.nlargest(datacoutt, 'Count1').reset_index()
    else:
        df = dfc[dfc['list_of_inventors'].isin(datacoui)]

    inventor_list = df['list_of_inventors'].tolist()
    year_list1 = df['Year'].tolist()
    co_list1 = df['Count1'].tolist()
    inventor_el = list()

    def first10(s):
        return s[:10]

    for i in inventor_list:
        str = i
        ten = first10(str)
        inventor_el.append(ten)

    n = len(inventor_el)
    for i in range(n):
        inventor_el[i] = inventor_el[i] + '...'
    if len(co_list1) == 0:
        annotations1 = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations1 = []
    if len(co_list1) == 0:
        annotations = [
            dict(

                text="No matching data found",
                showarrow=False,
                font=dict(
                    color="black",
                    size=34
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',

            )

        ]
    else:
        annotations = [
            dict(
                x=x,
                y=y,
                text='d' if y < 0 else z,  # Some conditional to define outliers

                showarrow=False,
                font=dict(
                    color="black",
                    size=14
                ),
                xanchor='center',  # Position of text relative to x axis (left/right/center)
                yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

            ) for x, y, z in zip(inventor_el, year_list1, co_list1)
        ]
    trace9 = go.Bar(x=inventor_el,
                    y=co_list1,
                    hovertext=inventor_list,
                    hoverinfo="text",
                    text=co_list1,
                    name='Bar',

                    textposition='auto'
                    )
    if len(co_list1) == 0:
        co_list1.append(0)
    trace10 = go.Scatter(
        x=inventor_el,
        y=year_list1,
        hovertext=inventor_list,
        hoverinfo="text",
        name='Bubble',
        visible=False,
        mode='markers',
        marker=dict(
            color=year_list1,
            size=co_list1,
            sizemode='area',
            sizeref=2. * max(co_list1) / (50. ** 2),
            sizemin=6,

            showscale=True
        )

    )
    datas = [trace9, trace10]
    updatemenus = [
        dict(
            active=0,
            buttons=list([
                dict(label="Bar",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "Top selected inventors",
                            'yaxis': {'title': 'No of patents', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations1}]),
                dict(label="Bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "Top selected inventors publication year wise",
                            'yaxis': {'title': 'Year', 'tickformat': ',d'},
                            'xaxis': {'title': 'Inventors'},
                            "annotations": annotations}]),

            ]),
            direction="down",
            # pad={"r": 10, "t": 10},
            showactive=True,
            x=0.1,
            xanchor="left",
            y=1.15,
            yanchor="top"
        )
    ]
    layout = dict(title='Top selected inventors', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False},
                  yaxis={'showgrid': False}, plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datas, layout=layout)
    divb13 = opy.plot(fig, auto_open=False, output_type='div')

    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    zippedList = list(zip(Country, Year))
    dfObj1x = pd.DataFrame(zippedList, columns=['Country', 'Year'])
    dfObj1x = dfObj1x.groupby(["Country", "Year"]).size().to_frame(name='Count1').reset_index()
    df = dfObj1x[(dfObj1x['Year'] >= datay) & (dfObj1x['Year'] <= datax)]
    dfc = df[df['Country'].isin(datacou)]
    df1 = dfc.pivot(index='Country', columns='Year', values='Count1')


    zx = dfc.pivot(index='Country', columns='Year', values='Count1').values

    zx[np.isnan(zx)] = 0
    axc = zx.tolist()

    yaxis = list(df1.index.values)
    xaxis = list(df1.columns)

    z11 = axc
    x11 = xaxis
    y11 = yaxis

    annotations = go.Annotations()
    for n, row in enumerate(z11):
        for m, val in enumerate(row):
            annotations.append(go.Annotation(text=z11[n][m], x=x11[m], y=y11[n],
                                             showarrow=False, font=dict(
                    color="black",
                    size=14
                )))
    if dfc.empty == True:
        trace00 = go.Bar(x=dfc['Year'],
                         y=dfc['Country'],
                         )
    else:
         trace00 = go.Heatmap(x=x11, y=y11, z=z11, colorscale='YlOrRd', showscale=False, name="heat", visible=False)
    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    g66 = list()
    for row in data:
        g66.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Year = list()
    for i in g66:
        datestring = i
        dt = try_parsing_date(datestring)
        Year.append(dt.year)

    zippedList = list(zip(Country, Year))
    dfObjp = pd.DataFrame(zippedList, columns=['Country', 'Year'])

    dfObjp = dfObjp.groupby(['Country', 'Year']).size().to_frame(name='Count').reset_index()
    dfObjp['size1'] = dfObjp.apply(lambda row: (row.Count * 1), axis=1)

    df1 = dfObjp[(dfObjp['Year'] >= datay) & (dfObjp['Year'] <= datax)]
    dfc = df1[df1['Country'].isin(datacou)]
    Country_list = dfc['Country'].tolist()
    Year_list = dfc['Year'].tolist()
    Count_list = dfc['Count'].tolist()
    size1_list = dfc['size1'].tolist()
    html_table1 = dfc.to_html(index=False)
    annotations1 = [
        dict(
            x=x,
            y=y,
            text='d' if x < 0 else z,  # Some conditional to define outliers

            showarrow=False,
            font=dict(
                color="black",
                size=14
            ),
            xanchor='center',  # Position of text relative to x axis (left/right/center)
            yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

        ) for x, y, z in zip(Year_list,Country_list,  Count_list)
    ]
    if len(Count_list) == 0:
        Count_list.append(0)
    trace11 = go.Scatter(
        x=Year_list,
        y=Country_list,
        name="bubble",
        mode='markers',
        marker=dict(
            color=Year_list,
            size=Count_list,
            sizemode='area',
            sizeref=3. * max(Count_list) / (30. ** 3),
            sizemin=6,

            showscale=True
        )

    )
    datase = [trace00, trace11]
    updatemenus = list([
        dict(
            active=1,
            buttons=list([
                dict(label="heat",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "No of patents of countries in respective years ",
                            'xaxis': {'title': 'Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations}]),
                dict(label="bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "No of patents of countries in respective years",
                            'xaxis': {'title': 'Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations1}]),

            ]),
        )
    ])
    layout = dict(title='No of patents of countries in respective years', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False}, yaxis={'showgrid': False},
                  plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=datase, layout=layout)
    divb14 = opy.plot(fig, auto_open=False, output_type='div')

    g33 = list()
    for row in data:
        g33.append(row.Publication_number)

    Country_ep = list()

    def first2(s):
        return s[:2]

    for i in g33:
        str = i
        two = first2(str)
        Country_ep.append(two)

    g44 = list()
    for row in data:
        g44.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g44:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    zippedList = list(zip(Country_ep, Earliest_year))
    dfObj1 = pd.DataFrame(zippedList, columns=['Country_ep', 'Earliest_year'])
    dfObj1 = dfObj1.groupby(["Country_ep", "Earliest_year"]).size().to_frame(name='Count1').reset_index()
    df3 = dfObj1[(dfObj1['Earliest_year'] >= datay) & (dfObj1['Earliest_year'] <= datax)]
    dfc = df3[df3['Country_ep'].isin(datacou)]
    df4 = dfc.pivot(index='Country_ep', columns='Earliest_year', values='Count1')
    z = dfc.pivot(index='Country_ep', columns='Earliest_year', values='Count1').values
    z[np.isnan(z)] = 0

    a = z.tolist()
    yaxis1 = list(df4.index.values)
    xaxis1 = list(df4.columns)
    z111 = a
    x111 = xaxis1
    y111 = yaxis1
    annotations = go.Annotations()
    for n, row in enumerate(z111):
        for m, val in enumerate(row):
            annotations.append(go.Annotation(text=z111[n][m], x=x111[m], y=y111[n],
                                             showarrow=False, font=dict(
                    color="black",
                    size=14
                )))
    if dfc.empty == True:
        trace000 = go.Bar(x=dfc['Earliest_year'],
                         y=dfc['Country_ep'],
                         )
    else:
        trace000 = go.Heatmap(x=x111, y=y111, z=z111, colorscale='YlOrRd', showscale=False, name="heat", visible=False)
    g77 = list()
    for row in data:
        g77.append(row.Publication_number)

    Country_ep = list()

    def first2(s):
        return s[:2]

    for i in g77:
        str = i
        two = first2(str)
        Country_ep.append(two)

    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)

    zippedList = list(zip(Country_ep, Earliest_year))
    dfObj4e = pd.DataFrame(zippedList, columns=['Country_ep', 'Earliest_year'])

    dfObj4e = dfObj4e.groupby(['Country_ep', 'Earliest_year']).size().to_frame(name='Count').reset_index()
    dfObj4e['size2'] = dfObj4e.apply(lambda row: (row.Count * 1), axis=1)
    df2 = dfObj4e[(dfObj4e['Earliest_year'] >= datay) & (dfObj4e['Earliest_year'] <= datax)]
    dfc = df2[df2['Country_ep'].isin(datacou)]

    Country_liste = dfc['Country_ep'].tolist()
    Year_liste = dfc['Earliest_year'].tolist()
    Count_lise = dfc['Count'].tolist()
    size1_liste = dfc['size2'].tolist()
    pd.set_option('display.width', 2000)
    html_table = dfc.to_html(index=False)


    annotations1 = [
        dict(
            x=x,
            y=y,
            text='d' if x < 0 else z,  # Some conditional to define outliers

            showarrow=False,
            font=dict(
                color="black",
                size=14
            ),
            xanchor='center',  # Position of text relative to x axis (left/right/center)
            yanchor='middle',  # Position of text relative to y axis (top/bottom/middle)

        ) for x, y, z in zip(Year_liste,Country_liste,  Count_lise)
    ]
    if len(Count_lise) == 0:
        Count_lise.append(0)
    trace111 = go.Scatter(
        x=Year_liste,
        y=Country_liste,
        name="bubble",
        mode='markers',
        marker=dict(
            color=Year_liste,
            size=Count_lise,
            sizemode='area',
            sizeref=3. * max(Count_lise) / (30. ** 3),
            sizemin=6,

            showscale=True
        )

    )
    dataset = [trace000, trace111]
    updatemenus = list([
        dict(
            active=1,
            buttons=list([
                dict(label="heat",
                     method="update",
                     args=[{"visible": [True, False]},
                           {"title": "No of patents of countries in respective priority years ",
                            'xaxis': {'title': 'priority Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations}]),
                dict(label="bubble",
                     method="update",
                     args=[{"visible": [False, True]},
                           {"title": "No of patents of countries in respective priority years",
                            'xaxis': {'title': 'priority Year', 'tickformat': ',d'},
                            'yaxis': {'title': 'Country'},
                            "annotations": annotations1}]),

            ]),
        )
    ])
    layout = dict(title='No of patents of countries in respective priority years', showlegend=False,
                  updatemenus=updatemenus, xaxis={'showgrid': False}, yaxis={'showgrid': False},
                  plot_bgcolor='rgb(252, 243, 207)')
    fig = go.Figure(data=dataset, layout=layout)
    divb15 = opy.plot(fig, auto_open=False, output_type='div')
    VBTl = VBT.tolist()
    zxyl = zxy.tolist()

    zctl = zct.tolist()
    zctcl = zctc.tolist()
    request.session['VBTl'] = VBTl
    request.session['zxyl'] = zxyl
    request.session['zz'] = zz
    request.session['zctl'] = zctl
    request.session['zctcl'] = zctcl
    request.session['div111'] = div111
    request.session['divb8'] = divb8
    request.session['divb10'] = divb10
    request.session['divb11'] = divb11
    request.session['divb12'] = divb12
    request.session['divb13'] = divb13
    request.session['divb14'] = divb14
    request.session['divb15'] = divb15
    return render(request, 'home1.html', {'z0t':VBTl,'z':zxyl,'z3':zz,'zs':zctl,'zs1':zctcl,'html_tablex':html_tablex})

def home1(request):
    data = models.report.objects.all()
    g55 = list()
    for row in data:
        g55.append(row.Publication_number)

    Country = list()

    def first2(s):
        return s[:2]

    for i in g55:
        str = i
        two = first2(str)
        Country.append(two)

    x = np.array(Country)
    z = (np.unique(x))

    zz = [10, 20, 30]
    s5 = list()
    for row in data:
        s5.append(row.Inventors)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_inventors = list()

    for sublist in aa5:
        for val in sublist:
            list_of_inventors.append(val)
    x = np.array(list_of_inventors)
    zc = (np.unique(x))

    s5 = list()
    for row in data:
        s5.append(row.Applicants)

    aa5 = list()
    for i in s5:
        str = i
        aa5.append(str.split('\n'))

    list_of_applicantsx = list()

    for sublist in aa5:
        for val in sublist:
            list_of_applicantsx.append(val)
    yt = np.array(list_of_applicantsx)
    yct = (np.unique(yt))

    g2 = list()
    for row in data:
        g2.append(row.Publication_date)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    a2 = list()
    for i in g2:
        datestring = i
        dt = try_parsing_date(datestring)
        a2.append(dt.year)
    x11 = np.array(a2)
    zxy1 = (np.unique(x11))

    g88 = list()
    for row in data:
        g88.append(row.Earliest_priority)

    def try_parsing_date(text):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                pass
        raise ValueError('no valid date format found')

    Earliest_year = list()
    for i in g88:
        datestring = i
        dt = try_parsing_date(datestring)
        Earliest_year.append(dt.year)
    yx11 = np.array(Earliest_year)
    yzxy1 = (np.unique(yx11))
    VB = np.insert(yzxy1, np.arange(len(zxy1)), zxy1)
    VBT = (np.unique(VB))

    html_tablex = request.session['html_tablex']
    return render(request, 'home1.html', {'z0t': VBT, 'z': z, 'z3': zz, 'zs': zc, 'zs1': yct,'html_tablex':html_tablex})
def home0(request):

    VBTl = request.session['VBTl']
    zxyl = request.session['zxyl']
    zz =request.session['zz']
    zctl =request.session['zctl']
    zctcl =request.session['zctcl']
    div111 = request.session['div111']
    divb8 = request.session['divb8']
    divb10 = request.session['divb10']
    divb11 = request.session['divb11']
    divb12 = request.session['divb12']
    divb13 = request.session['divb13']
    divb14 = request.session['divb14']
    divb15 = request.session['divb15']
    return render(request, 'home.html',{'z0t':VBTl,'z':zxyl,'z3':zz,'zs':zctl,'zs1':zctcl, 'plotdiv': div111,
                   'plotdiv9': divb8, 'plotdiv11': divb10,
                   'plotdiv12': divb11, 'plotdiv13': divb12, 'plotdiv14': divb13, 'plotdiv15': divb14,
                   'plotdiv16': divb15})